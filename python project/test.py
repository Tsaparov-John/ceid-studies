from openpyxl import Workbook, load_workbook
#import pandas as pd
from openpyxl.utils import get_column_letter
import numpy as np
from matplotlib import pyplot as plt
import mysql.connector
import csv
from matplotlib.ticker import ScalarFormatter




print("starts here")
# P GIA TA ARXEIA NSA K AAT
p=["nsa","aat"]   
for a in p:
    path=str(a)+".xlsx"
    wb=load_workbook(path)
    conn =  mysql.connector.connect(host="localhost", port="3306",user="root",password="",database="test")
    cursor = conn.cursor()
    print("i will ignore the warning  yolo \n")
    
    if a=="nsa":
        sheet=[1,2,3,7,8,9,13,14,15]
   
    if a=="aat":
        sheet=[1,2,3,7,8,9,13,14,15]
      
        sheet=[1,5,9]    
    for s in sheet:
    
        ws = wb.get_sheet_by_name("Sheet "+str(s))
        print(ws)
        table_name=ws['C8'].value
        print(table_name)
        name=str(a)+"pinakas"+str(s)
        print('---------------------')
        selectquery="CREATE TABLE "+name+" (Counter int AUTO_INCREMENT,TIME varchar(255),BG varchar(255) NULL,GR varchar(255) NULL, PRIMARY KEY (Counter));"
        print(selectquery)
        cursor.execute(selectquery)
        table = np.array(['TIME','BG','GR'])
        timecol=[]
        bgcol=[]
        grcol=[]
        for col in range (2,97,2):
            char=get_column_letter(col)

            time=ws[char+'10'].value
            bg=ws[char+'12'].value
            
            if bg==":" :
                bg="0"
            
            gr = ws[char+'13'].value

            timecol.append(time)
            bgcol.append(bg)
            grcol.append(gr)

            new_row = [ time,bg ,gr ]
            table = np.vstack([table,new_row])
            time="'"+time+"'"
            bg=str(bg)
            gr=str(gr)
           
            selectquery="INSERT INTO "+name+" (TIME, BG, GR) VALUES ("+time+","+bg+","+gr+");"
            cursor.execute(selectquery)
        print(table)
        
        plt.plot(timecol,grcol)
        plt.title(table_name, fontsize=6)
        plt.xlabel("time",fontsize=5)
        plt.ylabel("visitors",fontsize=5,rotation=90)
        plt.xticks(fontsize=5, rotation=90)
        plt.yticks(fontsize=5)
        
        
        file=name+"a.png"
        plt.savefig(file)
        plt.clf()



        plt.plot(timecol,bgcol)
        plt.title(table_name, fontsize=6)
        plt.xlabel("time",fontsize=5)
        plt.ylabel("visitors",fontsize=5,rotation=90)
        plt.xticks(fontsize=5, rotation=90)
        plt.yticks(fontsize=5)
        
        
        file=name+"b.png"
        plt.savefig(file)
        plt.clf()
        
        with open(name,'w') as f:
            write = csv.writer(f)
            write.writerows(table)


  
    
#NSB KAI AON
p=["nsb","aon"]   
for a in p:
    path=str(a)+".xlsx"
    wb=load_workbook(path)
    
    print("i will ignore the warning  yolo \n")
    
    if a=="nsb":
        sheet=[1]
   
    if a=="aon":
        sheet=[1,5,9]
      
        
    for s in sheet:
    
        ws = wb.get_sheet_by_name("Sheet "+str(s))
        print(ws)
        table_name=ws['C7'].value
        print(table_name)
        name=str(a)+"pinakas"+str(s)
        print('---------------------')
        selectquery="CREATE TABLE "+name+" (Counter int AUTO_INCREMENT,TIME varchar(255),BG varchar(255) NULL,GR varchar(255) NULL, PRIMARY KEY (Counter));"
        print(selectquery)
        cursor.execute(selectquery)
        table = np.array(['TIME','BG','GR'])
        timecol=[]
        bgcol=[]
        grcol=[]
        for col in range (2,97,2):
            char=get_column_letter(col)

            time=ws[char+'10'].value
            bg=ws[char+'12'].value
            
            if bg==":" :
                bg="0"
            
            gr = ws[char+'13'].value
            if gr==":" :
                gr="0"
            timecol.append(time)
            bgcol.append(bg)
            grcol.append(gr)

            new_row = [ time,bg ,gr ]
            table = np.vstack([table,new_row])
            time="'"+time+"'"
            bg=str(bg)
            gr=str(gr)
           
            selectquery="""INSERT INTO """+name+""" (TIME, BG, GR) VALUES ("""+time+""","""+bg+""","""+gr+""");"""
            cursor.execute(selectquery)
        print(table)
        
        plt.plot(timecol,grcol)
        plt.title(table_name, fontsize=6)
        plt.xlabel("time",fontsize=5)
        plt.ylabel("visitors",fontsize=5,rotation=90)
        plt.xticks(fontsize=5, rotation=90)
        plt.yticks(fontsize=5)
        
        
        file=name+"a.png"
        plt.savefig(file)
        plt.clf()



        plt.plot(timecol,bgcol)
        plt.title(table_name, fontsize=6)
        plt.xlabel("time",fontsize=5)
        plt.ylabel("visitors",fontsize=5,rotation=90)
        plt.xticks(fontsize=5, rotation=90)
        plt.yticks(fontsize=5)
        
        
        file=name+"b.png"
        plt.savefig(file)
        plt.clf()
        
        with open(name,'w') as f:
            write = csv.writer(f)
            write.writerows(table)


conn.commit()    
   
